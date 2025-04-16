import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side,Alignment
from openpyxl.worksheet.page import PageMargins
from pars import new_send_mail
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.pagebreak import ColBreak, RowBreak
from data import base_buratino, balance_buratino
from data import reset_the_balance
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl.drawing.image import Image
from email.mime.application import MIMEApplication


empty_shipping_positions = []


from openpyxl.drawing.text import Paragraph, CharacterProperties, RichTextProperties

def ask_for_category_with_dropdown(position, categories):#функция которая открывает диалоговое окно для пропущенных позиций
    # Создаем глобальную переменную для хранения результата
    selected_category = None

    # Функция для обработки выбора
    def on_submit():
        nonlocal selected_category
        selected_category = combo.get()  # Получаем выбранное значение
        root.destroy()  # Закрываем окно

    # Создаем окно
    root = tk.Tk()
    root.title("Выбор категории")
    root.geometry("500x150")
    #root.iconbitmap("1.ico")
    # Метка с текстом
    label = tk.Label(root, text=f"Выберите категорию для позиции: {position}")
    label.pack(pady=10)

    # Выпадающий список
    combo = ttk.Combobox(root, values=categories, state="readonly", width="50", height = "30")
    combo.pack(pady=5)
    combo.current(0)  # Устанавливаем первую категорию как выбранную по умолчанию

    # Кнопка подтверждения
    button = tk.Button(root, text="Подтвердить", command=on_submit)
    button.pack(pady=10)

    # Запуск главного цикла
    root.mainloop()

    return selected_category

def ask_for_category(position): #функция которая открывает диалоговое окно для пропущенных позиций
    root = tk.Tk()
    root.withdraw()  # Скрыть основное окно

    # Спрашиваем у пользователя категорию для позиции
    category = simpledialog.askstring("Категория", f"Введите категорию для позиции: {position}")

    return category

def cleaning(nametag):
    if not nametag:
        return ''
    nametag = nametag.replace(' ', '')
    nametag = nametag.replace('	', '')
    return nametag
    #return re.sub(r'[ \t]', '', nametag)

def create_a_table(yourname, date_table=''): #создает таблицу
    
    wb = Workbook()
    ws = wb.active

    alignment_style = Alignment(horizontal='center', vertical='center')
    border_style1 = Border(left=Side(border_style='medium', color='000000'), 
                        #right=Side(border_style='thin', color='000000'), 
                        top=Side(border_style='medium', color='000000'), 
                        bottom=Side(border_style='thin', color='000000'))
    border_style2 = Border(#left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='medium', color='000000'), 
                    bottom=Side(border_style='thin', color='000000'))
    
    border_style3 = Border(left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='thin', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='medium', color='000000'))
    
    border_style4 = Border(left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='thin', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='thin', color='000000'))
    
    border_style5 = Border(#left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='thin', color='000000'))
    
    border_style6 = Border(#left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='medium', color='000000'))
    
    border_style7 = Border(left=Side(border_style='medium', color='000000'), 
                    #right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='medium', color='000000'))
    
    border_style8 = Border(#left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='medium', color='000000'))
    
    border_style9 = Border(#left=Side(border_style='medium', color='000000'), 
                    #right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='thin', color='000000'))

    border_style10 = Border(#left=Side(border_style='medium', color='000000'), 
                    right=Side(border_style='medium', color='000000'), 
                    top=Side(border_style='thin', color='000000'), 
                    bottom=Side(border_style='thin', color='000000'))

    ws['B2'] = '30 КСЕРОКОПИЯ'
    ws['B2'].border  = border_style1
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.merge_cells('B2:C2')
    ws.merge_cells('D2:E2')
    ws['D2'] = yourname
    ws['D5'] = 'Распечатка: '
    ws['D2'].alignment = alignment_style
    ws['E2'].border  = border_style2
    ws['D2'].border  = border_style2

    ws['C5'] = 'Услуги: '
    ws['B8'] = 'Ламинир (120): '
    ws['B9'] = '(360)'
    ws['B10'] = '(750)'
    ws['C9'] = '(180)'
    ws['C10'] = '(850)'
    ws['B11'] = 'ЛАЗЕР: ' 
    ws['B14'] = 'А1(715)'
    ws['B15'] = 'А1(920): '
    ws['B16'] = 'А2(475)'
    ws['B17'] = 'А2(735)'
    ws['B18'] = 'Переплет: '

    ws['B21'] = 'Цв.ксер(45)'
    ws['B22'] = 'Цв.печать(40)'
    ws['B25'] = '10*15(40): '
    ws['B27'] = '(45): '
    ws['B28'] = '(50): '
    ws['B29'] = '(70): '
    ws['B30'] = '(100): '
    ws['C25'] = '13 на 18'

    ws['D9'] = 'Лишние'
    ws['D10'] = 'Каспи'
    ws['D11'] = 'Наличные '
    ws['D12'] = '1с: '
    ws['D13'] = 'ТОГО: '



    ws['D14'] = 'Флешки: '
    ws['D15'] = 'Тв.Переплет:'

    
    
    ws['D18'] = 'Прочее:'
    ws['D25'] = 'А4: '
    #обратная сторона
    ws['B31'] = 'Рулонная печать фотобумага : '
    ws['B32'] = 'А0+ (2500) : '
    ws['B33'] = 'А0 мат (2100) : '

    ws['B35'] = 'А1 м. плотный (1600) : '

    ws['B37'] = 'А0+ холст(5400) : '
    ws['B38'] = 'А0 холст(5000) : '
    ws['B39'] = 'А1 холст(4200) : '
 

    ws['B43'] = 'А1 м. тонк (1200) : '
   

    ws['C32'] = 'А0+ гл (2800) : '
    ws['C33'] = 'А0 гл (2500) : '
    ws['C34'] = 'А1 гл плот (2000) : '

    ws['C36'] = 'А1 сатин (2400) : '

    ws['C38'] = 'А0 самоклейка (3000) : '
    ws['C39'] = 'А1 самоклейка (2100) : '
    ws['C41'] = 'А1 гл тонк (1550) : '
 

    ws['D31'] = 'Рулонная печать простая бумага: '
    ws['D32'] = 'А0+  '
    ws['D33'] = 'А0  '
    ws['D34'] = 'А1  '
    ws['D35'] = 'А2  '
    ws['D36'] = 'А3  '

    ws['D37'] = 'УФ:  '
    ws['D38'] = 'Подрамники:  '
    ws['D39'] = 'Баннера:  '
    ws['D40'] = 'Фольга:  '

 


    
    dddatainpars = date_table
    ws['D3'] = f'Дата: {dddatainpars}'

    ws.merge_cells('B11:C11')
    ws.merge_cells('B12:C12')
    ws.merge_cells('B13:C13')
    ws.merge_cells('B8:C8')
    ws.merge_cells('B14:C14')
    ws.merge_cells('B15:C15')
    ws.merge_cells('B16:C16')
    ws.merge_cells('B17:C17')
    ws.merge_cells('B18:C18')
    ws.merge_cells('B19:C19')
    ws.merge_cells('B20:C20')
    ws.merge_cells('B21:C21')
    ws.merge_cells('B22:C22')
    ws.merge_cells('B23:C23')
    ws.merge_cells('B24:C24')
    ws.merge_cells('B26:C26')
    ws.merge_cells('B27:C27')
    ws.merge_cells('B28:C28')
    ws.merge_cells('B29:C29')
    ws.merge_cells('B30:C30')
    ws.merge_cells('B31:C31')

    ws.merge_cells('D31:E31')
    ws.merge_cells('D32:E32')
    ws.merge_cells('D33:E33')
    ws.merge_cells('D34:E34')
    ws.merge_cells('D35:E35')
    ws.merge_cells('D36:E36')
    

    ws['B7'].border = border_style3
    ws['B10'].border = border_style3 
    ws['B13'].border = border_style3
    ws['B17'].border = border_style3
    ws['B20'].border = border_style3
    ws['B24'].border = border_style3
    ws['B30'].border = border_style3

    ws['B3'].border = border_style4
    ws['B4'].border = border_style4
    ws['B5'].border = border_style4
    ws['B6'].border = border_style4
    ws['B8'].border = border_style4
    ws['B9'].border = border_style4
    ws['B11'].border = border_style4
    ws['B12'].border = border_style4
    ws['B14'].border = border_style4
    ws['B15'].border = border_style4
    ws['B16'].border = border_style4
    ws['B18'].border = border_style4
    ws['B19'].border = border_style4
    ws['B21'].border = border_style4
    ws['B22'].border = border_style4
    ws['B23'].border = border_style4
    ws['B25'].border = border_style4
    ws['B26'].border = border_style4
    ws['B27'].border = border_style4
    ws['B28'].border = border_style4
    ws['B29'].border = border_style4
    
    ws['C3'].border = border_style5
    ws['C4'].border = border_style5
    ws['C5'].border = border_style5
    ws['C6'].border = border_style5
    ws['C8'].border = border_style5
    ws['C9'].border = border_style5
    ws['C11'].border = border_style5
    ws['C12'].border = border_style5
    ws['C14'].border = border_style5
    ws['C15'].border = border_style5
    ws['C16'].border = border_style5
    ws['C18'].border = border_style5
    ws['C19'].border = border_style5
    ws['C21'].border = border_style5
    ws['C22'].border = border_style5
    ws['C23'].border = border_style5
    ws['C25'].border = border_style5
    ws['C26'].border = border_style5
    ws['C27'].border = border_style5
    ws['C28'].border = border_style5
    ws['C29'].border = border_style5

    
    ws['C7'].border = border_style6
    ws['C10'].border = border_style6
    ws['C13'].border = border_style6
    ws['C17'].border = border_style6
    ws['C20'].border = border_style6
    ws['C24'].border = border_style6
    ws['C30'].border = border_style6

    ws['D11'].border = border_style7
    ws['D13'].border = border_style7
    ws['D14'].border = border_style7
    ws['D24'].border = border_style7
    ws['D30'].border = border_style7


    ws['E11'].border = border_style8
    ws['E13'].border = border_style8
    ws['E14'].border = border_style8
    ws['E24'].border = border_style8
    ws['E30'].border = border_style8

    ws['D3'].border = border_style9
    ws['D4'].border = border_style9
    ws['D5'].border = border_style9
    ws['D6'].border = border_style9
    ws['D7'].border = border_style9
    ws['D8'].border = border_style9
    ws['D9'].border = border_style9
    ws['D10'].border = border_style9
    ws['D16'].border = border_style9
    ws['D17'].border = border_style9
    ws['D18'].border = border_style9
    ws['D19'].border = border_style9
    ws['D20'].border = border_style9
    ws['D21'].border = border_style9
    ws['D22'].border = border_style9
    ws['D23'].border = border_style9
    ws['D26'].border = border_style9
    ws['D27'].border = border_style9
    ws['D28'].border = border_style9
    ws['D29'].border = border_style9

    ws['E3'].border = border_style10
    ws['E4'].border = border_style10
    ws['E5'].border = border_style10
    ws['E6'].border = border_style10
    ws['E7'].border = border_style10
    ws['E8'].border = border_style10
    ws['E9'].border = border_style10
    ws['E10'].border = border_style10
    ws['E16'].border = border_style10
    ws['E17'].border = border_style10
    ws['E18'].border = border_style10
    ws['E19'].border = border_style10
    ws['E20'].border = border_style10
    ws['E21'].border = border_style10
    ws['E22'].border = border_style10
    ws['E23'].border = border_style10
    ws['E26'].border = border_style10
    ws['E27'].border = border_style10
    ws['E28'].border = border_style10
    ws['E29'].border = border_style10
    ws['E12'].border = border_style10
    ws['E15'].border = border_style10
    ws['E25'].border = border_style10

    ws['B31'].border = border_style4
    ws['B32'].border = border_style4
    ws['B33'].border = border_style4
    ws['B34'].border = border_style4
    ws['B35'].border = border_style4
    ws['B36'].border = border_style4
    ws['B37'].border = border_style4
    ws['B38'].border = border_style4
    ws['B39'].border = border_style4
    ws['B40'].border = border_style4
    ws['B41'].border = border_style4
    ws['B42'].border = border_style4
    ws['B43'].border = border_style4

    ws['B44'].border = border_style3

    ws['C32'].border = border_style5
    ws['C33'].border = border_style5
    ws['C34'].border = border_style5
    ws['C35'].border = border_style5
    ws['C36'].border = border_style5
    ws['C37'].border = border_style5
    ws['C38'].border = border_style5
    ws['C39'].border = border_style5
    ws['C40'].border = border_style5
    ws['C41'].border = border_style5
    ws['C42'].border = border_style5
    ws['C43'].border = border_style5

    ws['C44'].border = border_style6

    ws['D31'].border = border_style10
    ws['D32'].border = border_style10
    ws['D33'].border = border_style10
    ws['D34'].border = border_style10
    ws['D35'].border = border_style10

    ws['D36'].border = border_style8


    #настройка документа 
    cm = 0.5/2.54
    ws.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
    
    ws.print_area = 'A1:F60'

#col_breaks
    #col_number = 5  # the row that you want to insert page break
    #col_break = Break(id=col_number)  # create Break obj
    #ws.col_breaks.remove(col_break)  # insert page break



    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = True






    
    sava_name = (f'Рузльтат {date_table}.xlsx')
    wb.save(sava_name)

def unidentified_tag(sheete, num): #заполняем пустые ячейки
    categories = ["Распечатка А4 ЧБ",
                "Флешки",
                "Твердый Переплет",
                'Товары',
   
                'Фотобумага А4',
                'Фотобумага 100',
                'Фотобумага 70',
                'Фотобумага 50',
                'Фотобумага 45',
                'Фотобумага 40',
                'Фотобумага 13*18',
                'Цветная печать А4',
                'Цветная копия А4',
                'Переплет',
                'Ватман А1',
                'Ватман А2',
                'Лазер',
                'Ламинат',
                'Услуга',
                'Копия',

                'Инж А0+',
                'Инж А0',
                'Инж А1',
                'Инж А2',
                'Инж А3',

                'Холст А0+' ,
                'Холст А0' ,
                'Холст А1' ,

                'Самоклейка А0' ,
                'Самоклейка А1' ,

                'Сатин А1' ,

                'Мат тонкий А1' ,
                'Мат плотный А1' ,
                'Гл тонкий А1' ,
                'Гл плотный А1' ,

                'Мат А0' ,
                'Гл А0' ,

                'Гл А0+',
                'Мат А0+',

                'Подрамник',
                'УФ',
                'Фольга'
                ]
    
    for i in range(2, num):
        #print(sheete[i][2].value)
        if sheete[i][2].value == None:
            tegg = sheete[i][0].value

            
            empty_shipping_positions.append(tegg)
            print(f'НАДЕНА ПОЗИЦИЯ КОТОРАЯ НЕТ В БАЗЕ:  {tegg}')
            #sheete[i][2].value = ask_for_category(sheete[i][0].value)
            sheete[i][2].value = ask_for_category_with_dropdown(sheete[i][0].value, categories)

def new_send_mail(text, twoofile): # Отправка сообщений с файлом
    print('[x] Идет отправка файлов на почту оиждай \n')
    # Параметры отправки письма
    sender_email = 'andreisk06070136@yandex.ru'
    sender_password = 'fazaeabisnooemmo'
    receiver_email = 'andreisk06070136@yandex.ru'
    subject = 'Лог'
    message = text

    # Создаем письмо
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Добавляем текст в сообщении
    body = MIMEText(message)
    msg.attach(body)

    # Читаем файл Excel и добавляем его в письмо
    with open(twoofile, 'rb') as file:
        attachment = MIMEApplication(file.read(), _subtype='xlsx')
        attachment.add_header('Content-Disposition', 'attachment', filename=twoofile)
        msg.attach(attachment)

    # Отправляем письмо
    with smtplib.SMTP('smtp.yandex.ru', 587) as smtp:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

def assign_a_tag(sheete, num): # присвоить тег 
        #print(sheete['A1'].value)
        #print(num)
        sheete['C1'] = "Теги"
        for i in range(2, num):
            for f in base_buratino.keys():
                if cleaning(sheete[i][0].value) == cleaning(f):
                    #print('Совпадение найдено ')
                    sheete[i][2].value = base_buratino[f]
                three = sheete[i][0].value
                if three[:3] == 'Э/п':
                    sheete[i][2].value = 'Товары'
                if three[:21] == 'Папка-файл мультифора':
                    sheete[i][2].value = 'Товары'
                if three[:20] == 'Фотобумага X-GREE А4':
                    sheete[i][2].value = 'Фотобумага А4'
                if three[:20] == 'Фотобумага X-GREE A4':
                    sheete[i][2].value = 'Фотобумага А4'
                if three[:19] == 'Фотобумага KODAK A4':
                    sheete[i][2].value = 'Фотобумага А4'
                if three[:7] == 'Ламинат':
                    sheete[i][2].value = 'Ламинат'
                if three[:12] == 'Сканирование':
                    sheete[i][2].value = 'Услуга'
                if three[:5] == 'Лазер':
                    sheete[i][2].value = 'Лазер'
                if three[:9] == 'Ватман А1':
                    sheete[i][2].value = 'Ватман А1'
                if three[:9] == 'Ватман А2':
                    sheete[i][2].value = 'Ватман А1'
                if three[:8] == 'Пружинка':
                    sheete[i][2].value = 'Переплет'
                if three[:6] == 'Флешка':
                    sheete[i][2].value = 'Флешки'
                if three[:5] == 'Рамка':
                    sheete[i][2].value = 'Товары'
                if three[:9] == 'Подрамник':
                    sheete[i][2].value = 'Подрамник'
                if three[:11] == 'Аккумулятор':
                    sheete[i][2].value = 'Товары'
                if three[:15] == 'Папка с файлами':
                    sheete[i][2].value = 'Товары'
                if three[:10] == 'Фотоальбом':
                    sheete[i][2].value = 'Товары'
                if three[:6] == 'Краска':
                    sheete[i][2].value = 'Товары'
                if three[:10] == 'Обложка(2)':
                    sheete[i][2].value = 'Твердый Переплет'
                if three[:22] == 'Канал твердый переплет':
                    sheete[i][2].value = 'Твердый Переплет'
                if three[:10] == 'Брошюровка':
                    sheete[i][2].value = 'Переплет'
                

def counting_the_amounts(sheete, num): # считаем сумму
    reset_the_balance()

    for i in range(2, num):
        
        if i:
            for f in balance_buratino.keys():
                ci = cleaning(sheete[i][2].value)
                cf = cleaning(f)
                if ci == cf:
                    balance_buratino[f] = balance_buratino[f] + int(sheete[i][1].value)

def number_of_lines(fileName): #Считаем количество заполненных строк в таблице 
    book = openpyxl.open(fileName)
    sheete = book.active
    the_number_of_the_last_line = ''
    for i in range(1, 10000):
        #print(sheete[i][1].value)
        if  sheete[i][1].value == None:
            #print('PPPPPPPPPPPPPPPPP')
            the_number_of_the_last_line = i
            break
    return(the_number_of_the_last_line)

def writing_to_the_table(namee, jpgname): #запись в таблицу
    book = openpyxl.open(namee)
    sheete = book.active

    red_font = Font(color="FF0000", bold=True)
    


    for i in balance_buratino.keys():
        if balance_buratino[i] != 0:
            #print(balance_buratino[i])
            #print(i)
            if i == 'Распечатка А4 ЧБ':
                sheete['D5'] = f'{sheete['D5'].value} {balance_buratino[i]}'
                sheete['D5'].font = red_font
            elif i == 'Флешки':
                sheete['D14'] = f'{sheete['D14'].value} {balance_buratino[i]}'
                sheete['D14'].font = red_font
            elif i == 'Твердый Переплет':
                sheete['D15'] = f'{sheete['D15'].value} {balance_buratino[i]}'
                sheete['D15'].font = red_font
            elif i == 'Альбомы':
                sheete['D20'] = f'{sheete['D20'].value} {balance_buratino[i]}'
                sheete['D20'].font = red_font
            elif i == 'Рамки':
                sheete['D21'] = f'{sheete['D21'].value} {balance_buratino[i]}'
                sheete['D21'].font = red_font
            elif i == 'Скоросшиватель':
                sheete['D22'] = f'{sheete['D22'].value} {balance_buratino[i]}'
                sheete['D22'].font = red_font
            elif i == 'Папки':
                sheete['D23'] = f'{sheete['D23'].value} {balance_buratino[i]}'
                sheete['D23'].font = red_font
            elif i == 'Файлы':
                sheete['D24'] = f'{sheete['D24'].value} {balance_buratino[i]}'
                sheete['D24'].font = red_font
            elif i == 'Фотобумага А4':
                sheete['D25'] = f'{sheete['D25'].value} {balance_buratino[i]}'
                sheete['D25'].font = red_font
            elif i == 'Фотобумага 100':
                sheete['B30'] = f'{sheete['B30'].value} {balance_buratino[i]}'
                sheete['B30'].font = red_font
            elif i == 'Фотобумага 70':
                sheete['B29'] = f'{sheete['B29'].value} {balance_buratino[i]}'
                sheete['B29'].font = red_font
            elif i == 'Фотобумага 45':
                sheete['B27'] = f'{sheete['B27'].value} {balance_buratino[i]}'
                sheete['B27'].font = red_font
            elif i == 'Фотобумага 50':
                sheete['B28'] = f'{sheete['B28'].value} {balance_buratino[i]}'
                sheete['B28'].font = red_font
            elif i == 'Фотобумага 40':
                sheete['B25'] = f'{sheete['B25'].value} {balance_buratino[i]}'
                sheete['B25'].font = red_font
            elif i == 'Цветная печать А4':
                sheete['B22'] = f'{sheete['B22'].value} {balance_buratino[i]}'
                sheete['B22'].font = red_font
            elif i == 'Цветная копия А4':
                sheete['B21'] = f'{sheete['B21'].value} {balance_buratino[i]}'
                sheete['B21'].font = red_font
            elif i == 'Переплет':
                sheete['B18'] = f'{sheete['B18'].value} {balance_buratino[i]}'
                sheete['B18'].font = red_font
            elif i == 'Ватман А1':
                sheete['B15'] = f'{sheete['B15'].value} {balance_buratino[i]}'
                sheete['B15'].font = red_font
            elif i == 'Ватман А2':
                sheete['B17'] = f'{sheete['B17'].value} {balance_buratino[i]}'
                sheete['B17'].font = red_font
            elif i == 'Лазер':
                sheete['B11'] = f'{sheete['B11'].value} {balance_buratino[i]}'
                sheete['B11'].font = red_font
            elif i == 'Ламинат':
                sheete['B8'] = f'{sheete['B8'].value} {balance_buratino[i]}'
                sheete['B8'].font = red_font
            elif i == 'Услуга':
                sheete['C5'] = f'{sheete['C5'].value} {balance_buratino[i]}'
                sheete['C5'].font = red_font
            elif i == 'Копия':
                sheete['B3'] = f'{balance_buratino[i]}'
                sheete['B3'].font = red_font
            elif i == 'Инж А0+':
                sheete['D32'] = f'{sheete['D32'].value} {balance_buratino[i]}'
                sheete['D32'].font = red_font
            elif i == 'Инж А0':
                sheete['D33'] = f'{sheete['D33'].value} {balance_buratino[i]}'
                sheete['D33'].font = red_font
            elif i == 'Инж А1':
                sheete['D34'] = f'{sheete['D34'].value} {balance_buratino[i]}'
                sheete['D34'].font = red_font
            elif i == 'Инж А2':
                sheete['D35'] = f'{sheete['D35'].value} {balance_buratino[i]}'
                sheete['D35'].font = red_font
            elif i == 'Инж А3':
                sheete['D36'] = f'{sheete['D36'].value} {balance_buratino[i]}'
                sheete['D36'].font = red_font
            elif i == 'Батарейки':
                sheete['D19'] = f'{sheete['D19'].value} {balance_buratino[i]}'
                sheete['D19'].font = red_font


            elif i == 'Холст А0+':
                sheete['B37'] = f'{sheete['B37'].value} {balance_buratino[i]}'
                sheete['B37'].font = red_font
            elif i == 'Холст А0':
                sheete['B38'] = f'{sheete['B38'].value} {balance_buratino[i]}'
                sheete['B38'].font = red_font
            elif i == 'Холст А1':
                sheete['B39'] = f'{sheete['B39'].value} {balance_buratino[i]}'
                sheete['B39'].font = red_font

            elif i == 'Самоклейка А0':
                sheete['C38'] = f'{sheete['C38'].value} {balance_buratino[i]}'
                sheete['C38'].font = red_font
            elif i == 'Самоклейка А1':
                sheete['C39'] = f'{sheete['C39'].value} {balance_buratino[i]}'
                sheete['C39'].font = red_font

            elif i == 'Сатин А1':
                sheete['C36'] = f'{sheete['C36'].value} {balance_buratino[i]}'
                sheete['C36'].font = red_font

            elif i == 'Мат тонкий А1':
                sheete['B43'] = f'{sheete['B43'].value} {balance_buratino[i]}'
                sheete['B43'].font = red_font
            elif i == 'Мат плотный А1':
                sheete['B35'] = f'{sheete['B35'].value} {balance_buratino[i]}'
                sheete['B35'].font = red_font
            elif i == 'Гл тонкий А1':
                sheete['C41'] = f'{sheete['C41'].value} {balance_buratino[i]}'
                sheete['C41'].font = red_font
            elif i == 'Гл плотный А1':
                sheete['C34'] = f'{sheete['C34'].value} {balance_buratino[i]}'
                sheete['C34'].font = red_font

            elif i == 'Мат А0':
                sheete['B33'] = f'{sheete['B33'].value} {balance_buratino[i]}'
                sheete['B33'].font = red_font
            elif i == 'Гл А0':
                sheete['C33'] = f'{sheete['C33'].value} {balance_buratino[i]}'
                sheete['C33'].font = red_font
            elif i == 'Гл А0+':
                sheete['C32'] = f'{sheete['C32'].value} {balance_buratino[i]}'
                sheete['C32'].font = red_font
            elif i == 'Мат А0+':
                sheete['B32'] = f'{sheete['B32'].value} {balance_buratino[i]}'
                sheete['B32'].font = red_font

            elif i == 'Товары':
                sheete['D18'] = f'{sheete['D18'].value} {balance_buratino[i]}'
                sheete['D18'].font = red_font

            elif i == 'Фотобумага 13*18':
                sheete['C25'] = f'{sheete['C25'].value} {balance_buratino[i]}'
                sheete['C25'].font = red_font

            elif i == 'Подрамник':
                sheete['D38'] = f'{sheete['D38'].value} {balance_buratino[i]}'
                sheete['D38'].font = red_font

            elif i == 'УФ':
                sheete['D37'] = f'{sheete['D37'].value} {balance_buratino[i]}'
                sheete['D37'].font = red_font

            elif i == 'Фольга':
                sheete['D40'] = f'{sheete['D40'].value} {balance_buratino[i]}'
                sheete['D40'].font = red_font

    iiii  = 1  
    for f in balance_buratino.keys():
        iiii = iiii + 1
        sheete[f'G{iiii}'] = f
    iiii  = 1  
    for f in balance_buratino.values():
        iiii = iiii + 1
        sheete[f'H{iiii}'] = f

    summ_oenc = 0   

    for f in balance_buratino.keys():
        summ_oenc = summ_oenc + balance_buratino[f]
        #print(summ_oenc)

    sheete['D12'] = f'СУММА 1С:  {summ_oenc}'
    sheete['D12'].font = red_font



    img(jpgname, sheete)
    book.save(namee)

def img(nameimage, sheete):
    if nameimage != 'Название фотки ':
        img = Image(nameimage)
        sheete.add_image(img, 'D41')
        img.width = 350 
        img.height = 400


def pars_buratino(name, date,  nameinput, jpgname): #основной цикл
    #fileName = input('Введи название таблицы:  ') + '.xlsx'
    fileName = name
    date_table = date
    sava_name = (f'Рузльтат {date_table}.xlsx')

    nameuser =  nameinput

    book = openpyxl.open(fileName)
    sheete = book.active

    num = number_of_lines(fileName)

    assign_a_tag(sheete, num) #присвоить тег 

    unidentified_tag(sheete, num) #найтим пустые ячейки и присвоить тег вручную

    counting_the_amounts(sheete, num)
    book.save(fileName)

    create_a_table(nameuser, date_table)

    writing_to_the_table(sava_name, jpgname)

    new_send_mail(str(empty_shipping_positions), sava_name)




