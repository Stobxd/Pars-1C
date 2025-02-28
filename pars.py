from socket import setdefaulttimeout
import wmi
import openpyxl
import os.path
import data
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from colorama import Fore, Back, Style
import smtplib
import tkinter as tk
from tkinter import ttk
#from buratino_mode import pars_buratino
import webbrowser
from data import vetsion


#pyinstaller --onefile --console --icon=1.ico pars.py
base = data.base
balance = data.balance
reset_the_balance = data.reset_the_balance
dataForTransportation = data.dataForTransportation

#print('[x] ВВЕДИ НАЗВАНИЕ ТОЧКИ')
#userpk = str(input())
#print(f'Спасибо {userpk}\n')
empty_shipping_positions = []

def ask_for_category_with_dropdown(position, categories):#функция которая открывает диалоговое окно для пропущенных позиций
    # Создаем глобальную переменную для хранения результата
    selected_category = None

    # Функция для обработки выбора
    def on_submit():
        nonlocal selected_category
        selected_category = combo.get()  # Получаем выбранное значение
        root.destroy()  # Закрываем окно

    # Создаем окно
    root = tk.Tk()
    root.title("Выбор категории")
    root.geometry("500x150")
    #root.iconbitmap("1.ico")
    # Метка с текстом
    label = tk.Label(root, text=f"Выберите категорию для позиции: {position}")
    label.pack(pady=10)

    # Выпадающий список
    combo = ttk.Combobox(root, values=categories, state="readonly", width="50", height = "30")
    combo.pack(pady=5)
    combo.current(0)  # Устанавливаем первую категорию как выбранную по умолчанию

    # Кнопка подтверждения
    button = tk.Button(root, text="Подтвердить", command=on_submit)
    button.pack(pady=10)

    # Запуск главного цикла
    root.mainloop()

    return selected_category

def unidentified_tag(sheete, num): #заполняем пустые ячейки
    categories = ['Товар (продажа)',
    'Обычная бумага А4 (ч/б печ)',
    'Обычная бумага А4 (ч/б ксер)',
    'Обычная бумага А4 (цв ксер/печ)',
    'Обычная бумага А3 (ч/б)',
    'Обычная бумага А3 (цвет)',
    'Фотобумага (40)',
    'Фотобумага (45)',
    'Фотобумага (50)',
    'Фотобумага (70)',
    'Фотобумага (100)',
    'Фотобумага (13 на 18)',
    'Фотобумага (а5/А4)',
    'Фотобумага (А3)',
    'Широкоформатка (Холст)',
    'Широкоформатка (Ватман А1, А2)',
    'Широкоформатка (Фото бум)',
    'Ламинат',
    'Фото на док',
    'Фотошоп',
    'Услуги (прочее)',
    'Лазер'
                ]
    
    for i in range(2, num):
        #print(sheete[i][2].value)
        if sheete[i][2].value == None:
            tegg = sheete[i][0].value

            
            empty_shipping_positions.append(tegg)
            print(f'НАДЕНА ПОЗИЦИЯ КОТОРАЯ НЕТ В БАЗЕ:  {tegg}')
            #sheete[i][2].value = ask_for_category(sheete[i][0].value)
            sheete[i][2].value = ask_for_category_with_dropdown(sheete[i][0].value, categories)
    
def version() :
    ver = vetsion
    return(ver)

def mistakes(twoofile) : # Создаем лог ошибки
    numberoflines = 1 
    booke = openpyxl.open(twoofile)
    sheete = booke.active 
    for row in range(1, 100):
        number = sheete[row][0].value
        if number:
            numberoflines = numberoflines + 1
    listtoandrew = []
    for row in range(1, numberoflines):
        number = sheete[row][2].value
        if number == None:
            number3 = sheete[row][0].value
            listtoandrew.append(number3)
        print(number)
    print('\n[x] Данные позиции были пропущены: ' + str(listtoandrew) + '\n')
    print('''[x] Возможно некоторые позиции были пропущены и
    теперь я хочу их передать на почту чтобы потом добавить в спиоск  \n''')

    print('[x] Идет отправка на почту оиждай \n')
    text = str(listtoandrew)
    # send_mails(text + '\nHWID: ' + defender())
    new_send_mail(text + '\n' + version() + ' Пользователь: ' + defender() + '\n' + 'User name: ', twoofile)
    print('[x] Отправка завершена \n')


    


    unidentified_tag(sheete, numberoflines)
    booke.save(twoofile)

def defender() : # Защита по HWID
    c = wmi.WMI()
    # listofavailableidyshniks = ['           Z7R4P3AAT', 'S4BFNJ0MB40846P']
    seriesofthisPC = []

    for item in c.Win32_PhysicalMedia():
        seriall = item.SerialNumber
        seriesofthisPC.append(str(seriall).replace(' ', ''))

    HWIDSTR = str(seriesofthisPC)

    for item in seriesofthisPC:
        if item == 'Z7R4P3AAT':
            HWIDSTR = 'Мой пк'
            break
        if item == 'B313079316EC00032665':
            HWIDSTR = 'Караван 1 ПК'
            break
        if item == 'Y3J33A6AS':
            HWIDSTR = 'Букетова'
            break
        if item == 'J32655J003368':
            HWIDSTR = 'Рынок'
            break
        if item == 'S4Y0S4P7':
            HWIDSTR = 'Океан'
            break
        if item == 'X7IYT1BHT':
            HWIDSTR = 'Караван'
            break
    return(HWIDSTR)

def choosingAnAction() : # Меню 

    print(version(),
        '''


Выбери операцию: 
    [x] Присвоить тег (СТАРОЕ)
    [x] Найти суммы (СТАРОЕ)

    [3] Две таблицы
    [4] Одна таблица
    [5] Таблица для буратино

    [6] Перейти на сайт для загрузки новой версии 
    ''')
    key = input('ПРОСТО ВВЕДИ ЧИСЛО: ')
    if key == "1":
        comparison(input('Введи название файла: ') + '.xlsx')
        print("\n\n\nЭта хуйня завершила свою работу возможно (проверок нет)")
        choosingAnAction()
    elif key == "2":
        theAmount(input('Введи название файла: ') + '.xlsx')
        print("\n\n\nЭта хуйня завершила свою работу возможно (проверок нет)")
        choosingAnAction()
    elif key == "3":
        merging_files()
        print("\n\n\nЭта хуйня завершила свою работу возможно (проверок нет)")
        choosingAnAction()
    elif key == "4":
        solo_mode()
        print("\n\n\nЭта хуйня завершила свою работу возможно (проверок нет)")
        choosingAnAction()
    elif key == "5":
        #pars_buratino()
        choosingAnAction()
    elif key == "6":
        
        url = 'https://github.com/Stobxd/Pars-1C/releases'
        webbrowser.open(url)    
            
        choosingAnAction()

    else :
        choosingAnAction()

def comparison(fileName) : # Присвоить тег
    number_of_matches = 0 #Считаем количество строк
    fileName

    check_for_file_existence (fileName) #проверка на наличие файла

    booke = openpyxl.open(fileName)
    sheete = booke.active 
    cellvalue = sheete['C1'].value

    if cellvalue:
        print('Столбец создан')
    else:
        sheete['C1'] = "Теги"


    row_count = sheete.max_row + 1
    # основной цикл 
    for row in range(1, row_count) :
        number = sheete[row][0].value    
        if number:
            Number2 = garbage_cleaning(number)
            for f in base.keys():
                # основной цикл 
                if Number2 == garbage_cleaning(f):
                    sheete[row][2].value = base[f]
                    number_of_matches = number_of_matches + 1
                    break
                #универсальные фильтры Рамка
                three = sheete[row][0].value
                if three[:3] == 'Э/п':
                    sheete[row][2].value = 'Товар (продажа)'
                if three[:6] == 'Флешка':
                    sheete[row][2].value = 'Товар (продажа)'
                if three[:7] == 'Ламинат':
                    sheete[row][2].value = 'Ламинат'
                if three[:5] == 'Рамка':
                    sheete[row][2].value = 'Товар (продажа)'

    print('\n\n\n[x] Найдено совпадений: '+ str(number_of_matches) + '\n') 

    # savename = input("\n\nВведи название файла для сохраниения: ")
    booke.save(fileName)
    twoofile = fileName
    mistakes(twoofile)
    #os.system("PAUSE")

def theAmount(fileName) : # Считаем сумму
    fileName
    check_for_file_existence (fileName)
    booke = openpyxl.open(fileName)
    sheete = booke.active 
    cellvalue = sheete['D1'].value

    if cellvalue:
        print('Столбец создан')
    else:
        sheete['D1'] = 'Отчет о проделанной работе'
        print('Создаем стобец')

    reset_the_balance()
    row_count = sheete.max_row + 1
    for row in range(1, row_count) :
        number = sheete[row][2].value
        if number:
            numberr = garbage_cleaning(number)
            for i in balance.keys():
                ii = garbage_cleaning(i)
                if numberr == ii:
                    balance[i] = balance[i] + int(sheete[row][1].value)

    # Считаем количество заполненых строк и записыавем в numberoflines


    # Делаем проверку на значения которые не были учтены в цикле 
    for row in range(1, row_count):
        cellkey = sheete[row][2].value
        if cellkey:
            cellkey = garbage_cleaning(cellkey)
            for i in balance.keys():
                ii = garbage_cleaning(i)
                if cellkey == ii:
                    sheete[row][3].value = 'Программа посчитала эту хуйню'
    

    # Запись в файл 
    file = open('Лог.txt', 'a', encoding='utf-8')
    for i in balance.keys():
        print(i + ' : ' + str(balance[i]))
        file.write(i + ' : ' + str(balance[i]) + '\n')
    
    #считаем сумму строк которые учли 
    summi = 0
    for i in balance:
        summi = summi + balance[i]
    print('Итог: ' + str(summi))
    file.write('\n' + 'Данные которые учел скрипт: ' + str(summi) + '\n')

    #Считаем обущую сумму екселя
    print("Считаем общую сумму екселя")
    summxlsx = 0
    for i in range(2, row_count):
        print(sheete[i][1].value)
        summxlsx = summxlsx + sheete[i][1].value
    print(summxlsx)
    file.write('Общая сумма екселя: ' + str(summxlsx) + '\n')

    #Находим разницу
    differenceofamounts = summxlsx - summi
    file.write('Разница: ' + str(differenceofamounts) + '\n \n')

    booke.save(fileName)
    #os.system("PAUSE")

def check_for_file_existence (fileName) : # Проверка наличия файла
    filesearch = os.path.exists(fileName)
    if filesearch == False:
        print(Fore.RED + Back.BLACK + '''
        [ОШИБКА] КРЧ НЕПРАВИЛЬНО ВВЕЛ ДАННЫЕ ВВЕДИ ЕЩЕ РАЗ !!!        
''' + Style.RESET_ALL)
        os.system("PAUSE")
        choosingAnAction()

def new_send_mail(text, twoofile): # Отправка сообщений с файлом
    print('[x] Идет отправка файлов на почту оиждай \n')
    # Параметры отправки письма
    sender_email = 'andreisk06070136@yandex.ru'
    sender_password = 'fazaeabisnooemmo'
    receiver_email = 'andreisk06070136@yandex.ru'
    subject = 'Лог'
    message = text

    # Создаем письмо
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Добавляем текст в сообщении
    body = MIMEText(message)
    msg.attach(body)

    # Читаем файл Excel и добавляем его в письмо
    with open(twoofile, 'rb') as file:
        attachment = MIMEApplication(file.read(), _subtype='xlsx')
        attachment.add_header('Content-Disposition', 'attachment', filename=twoofile)
        msg.attach(attachment)

    # Отправляем письмо
    with smtplib.SMTP('smtp.yandex.ru', 587) as smtp:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

def merging_files() : # Автоматический режим работы 
    fileNames = []
    fileName = input("Введи название двух файлов через запятую: ")
    fileNames = fileName.split(", ")
    theFirstTable = fileNames[0] + ".xlsx"
    theSecondTable = fileNames[1] + ".xlsx"
    
    book = openpyxl.open(theFirstTable)
    page = book.active
   
    numberofrows = line_count(theFirstTable) + 1
    
    hghg = []
    fdgfdge = []

    for i in range(2, numberofrows):
        hghg.append(page[i][0].value)

    for i in range(2, numberofrows):
        fdgfdge.append(page[i][1].value)

    book2 = openpyxl.open(theSecondTable)
    page2 = book2.active

    numberofrows2 = 0
    for i in page2:
        numberofrows2 +=1

    s = numberofrows2
    for i in hghg:
        s = s + 1
        page2[s][0].value = i

    s2 = numberofrows2
    for i in fdgfdge:
        s2 = s2 + 1
        page2[s2][1].value = i

    doublesavename = 'РЕЗУЛЬТАТ_' + fileNames[0] + fileNames[1] + '.xlsx'
    book2.save(doublesavename)
    comparison(doublesavename)
    theAmount(doublesavename)

    choosingAnAction()

def garbage_cleaning(garbage): # очистка мусора
    garbage = garbage.replace(' ', '')
    garbage = garbage.replace('	', '')
    return garbage

def countthenumberoflines(nametable): # тестовая залупа для просчета количества строк 
    numberoflines = 0
    for i in nametable:
        numberoflines += 1
        return numberoflines

def line_count(name): #считаем количество строк
    if type(name) == int:
        name = str(name)
    if name[-5:] != ".xlsx":
        name = str(name) + ".xlsx"
    check_for_file_existence(name)
    book = openpyxl.open(name)
    page = book.active


    row_count = page.max_row + 1
    #count = 0
    #for item in page:
    #    count += 1
    return row_count

def solo_mode(name, date) : # Solo mode 
    fileName = name
    print(fileName)
    
    book = openpyxl.open(fileName)
    page = book.active


    doublesavename = 'РЕЗУЛЬТАТ' + date + '.xlsx'
    book.save(doublesavename)
    comparison(doublesavename)
    theAmount(doublesavename)

  



#choosingAnAction()

#os.system("PAUSE")
            